"""Annotate the leftover-inventory lists with the SKU each part belongs to.

Reads the raw inventory workbooks in inputs/inventory_lists/, inserts three columns
(color / SKU / SKU description) after the part name, adds a mapping-basis notes column
and a legend, and writes a second sheet estimating how many units of each SKU can still
be built from the remaining stock.

The colour -> SKU rules live in inputs/mappings/*.csv, not in this file, so they can be
corrected without touching code.

Usage:
    python map_inventory_skus.py                 # both families
    python map_inventory_skus.py --family lsp    # one family
"""

from __future__ import annotations

import argparse
import csv
import math
import re
import sys
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.formula.translate import Translator
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
INPUTS = ROOT / "inputs"
MAPPINGS = INPUTS / "mappings"
OUTPUTS = ROOT / "outputs"

# --- source sheet layout (identical in both inventory workbooks) -----------------
HEADER_ROW = 3
FIRST_DATA_ROW = 4
COL_ITEM = 1
COL_PART_NAME = 2
COL_QTY_LOMAK = 3  # before the insert
N_INSERTED = 3     # color, SKU, SKU description

# column positions after the 3-column insert
COL_COLOR = 3
COL_SKU = 4
COL_SKU_DESC = 5
COL_COST_RMB = 10
COL_COST_HKD = 12
COL_COST_USD = 14
COL_NOTES = 23  # W, with V left blank as a spacer

FX = {"rmb": 6.8, "hkd": 7.8317, "usd": 1.0}

FILL_INFERRED = PatternFill("solid", fgColor="FFF2CC")
FILL_REVIEW = PatternFill("solid", fgColor="FCE4D6")

# One typeface for the whole output: Arial 9, black. Bold is the only variation.
FONT_NAME = "Arial"
FONT_SIZE = 9
FONT_COLOR = "FF000000"
BODY_FONT = Font(name=FONT_NAME, size=FONT_SIZE, color=FONT_COLOR)
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, color=FONT_COLOR, bold=True)
TITLE_FONT = HEADER_FONT
SECTION_FONT = HEADER_FONT

# Misspellings in the source lists that would otherwise hide a colour token.
SPELLING_ALIASES = {"burshed": "brushed", "seelve": "sleeve"}

NOTE_PREFIX = {
    "verified": "Verified",
    "inferred": "Inferred",
    "review": "NEEDS REVIEW",
}
REVIEW_DESC = "allocation not determinable from plytix - needs confirmation"


@dataclass
class Family:
    key: str
    source: Path
    sku_csv: Path
    color_map_csv: Path
    overrides_csv: Path
    output: Path
    scope: str
    extra_legend: list = field(default_factory=list)


FAMILIES = {
    "lsp": Family(
        key="lsp",
        source=INPUTS / "inventory_lists" / "lsp_inventory_list.xlsx",
        sku_csv=INPUTS / "lsp_skus.csv",
        color_map_csv=MAPPINGS / "lsp_color_map.csv",
        overrides_csv=MAPPINGS / "lsp_overrides.csv",
        output=OUTPUTS / "lsp_inventory_sku_mapping.xlsx",
        scope='Scope: this list covers the 6-SKU "9 oz. sensor pump" family only.',
        extra_legend=[
            "ST1092 (polished): no part on this list carries a polished finish, so no "
            "ST1092 units can be built from this stock.",
        ],
    ),
    "mirrors": Family(
        key="mirrors",
        source=INPUTS / "inventory_lists" / "sensor_mirror_inventory_list.xlsx",
        sku_csv=INPUTS / "mirrors_skus.csv",
        color_map_csv=MAPPINGS / "mirrors_color_map.csv",
        overrides_csv=MAPPINGS / "mirrors_overrides.csv",
        output=OUTPUTS / "sensor_mirror_inventory_sku_mapping.xlsx",
        scope='Scope: this list covers the 6-SKU "sensor mirror with touch-control '
              'brightness and dual light setting" family only.',
    ),
}


# --------------------------------------------------------------------------------
# mapping inputs
# --------------------------------------------------------------------------------
def read_skus(path):
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return {r["SKU"].strip(): r["Label"].strip() for r in csv.DictReader(fh)}


@dataclass
class ColorRule:
    color: str
    skus: str
    confidence: str
    note: str


def read_color_map(path):
    with path.open(newline="", encoding="utf-8-sig") as fh:
        rules = [
            ColorRule(
                (r["color"] or "").strip().lower(),
                (r["skus"] or "").strip(),
                (r["confidence"] or "inferred").strip().lower(),
                (r["note"] or "").strip(),
            )
            for r in csv.DictReader(fh)
        ]
    # Longest colour token first so "matte black" wins over "black".
    return sorted(rules, key=lambda r: len(r.color), reverse=True)


def read_overrides(path):
    if not path.exists():
        return {}
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return {
            int(r["item"]): r
            for r in csv.DictReader(fh)
            if (r.get("item") or "").strip()
        }


# --------------------------------------------------------------------------------
# per-part resolution
# --------------------------------------------------------------------------------
@dataclass
class Part:
    row: int
    item: int
    name: str
    color: str
    skus: list          # [] for ALL / REVIEW
    sku_text: str       # "ST3052 / ST3053", "ALL" or "REVIEW"
    description: str
    confidence: str
    note: str
    qty: float
    qty_per_unit: float  # None => pieces-per-unit unknown
    unit_cost_usd: float  # None => no cost on the list


def normalise(name):
    text = name.lower()
    for wrong, right in SPELLING_ALIASES.items():
        text = text.replace(wrong, right)
    return text


def detect_color(name, rules):
    text = normalise(name)
    for rule in rules:
        if rule.color and re.search(r"\b" + re.escape(rule.color) + r"\b", text):
            return rule
    return None


def default_rule(rules):
    for rule in rules:
        if not rule.color:
            return rule
    return ColorRule("", "ALL", "inferred", "")


def describe(skus, labels):
    """Collapse several SKU labels into one line by factoring out the shared text.

    ["...dual light setting, brushed steel", "...dual light setting, brass steel"]
      -> "...dual light setting, brushed / brass steel"
    """
    texts = [labels[s] for s in skus if s in labels]
    if not texts:
        return ""
    if len(texts) == 1:
        return texts[0]
    prefix = texts[0]
    for other in texts[1:]:
        while not other.startswith(prefix):
            prefix = prefix[:-1]
    suffix = texts[0][len(prefix):]
    for other in texts[1:]:
        tail = other[len(prefix):]
        while suffix and not tail.endswith(suffix):
            suffix = suffix[1:]
    cores = [t[len(prefix): len(t) - len(suffix)] if suffix else t[len(prefix):] for t in texts]
    return prefix + " / ".join(cores) + suffix


def unit_cost_usd(ws, row):
    for col, rate in (
        (COL_COST_RMB, FX["rmb"]),
        (COL_COST_HKD, FX["hkd"]),
        (COL_COST_USD, FX["usd"]),
    ):
        value = ws.cell(row, col).value
        if isinstance(value, (int, float)) and value:
            return float(value) / rate
    return None


def resolve_part(ws, row, rules, overrides, labels, sku_colors, warnings):
    item = int(ws.cell(row, COL_ITEM).value)
    name = str(ws.cell(row, COL_PART_NAME).value).strip()

    # A part naming its own SKU ("Gift box ST3052") outranks colour detection.
    stated = [s for s in labels if re.search(r"\b" + re.escape(s) + r"\b", name, re.I)]
    if stated:
        color = sku_colors.get(stated[0], "n/a")
        skus_text = " / ".join(stated)
        confidence = "verified"
        note = "colour named in part, matched to plytix label - SKU stated in part name"
    else:
        rule = detect_color(name, rules) or default_rule(rules)
        color = rule.color or "n/a"
        skus_text, confidence, note = rule.skus, rule.confidence, rule.note

    qty_per_unit = 1.0
    override = overrides.get(item)
    if override:
        expected = (override.get("part_name") or "").strip()
        if expected and expected.lower() != name.lower():
            warnings.append(
                "override for item %d expects part '%s' but the list now says '%s'"
                % (item, expected, name)
            )
        if (override.get("skus") or "").strip():
            skus_text = override["skus"].strip()
        if (override.get("confidence") or "").strip():
            confidence = override["confidence"].strip().lower()
        if (override.get("note") or "").strip():
            note = override["note"].strip()
        raw_qpu = (override.get("qty_per_unit") or "").strip()
        if raw_qpu == "?":
            qty_per_unit = None
        elif raw_qpu:
            qty_per_unit = float(raw_qpu)

    skus = [s.strip() for s in skus_text.split("/") if s.strip() in labels]
    if skus_text.upper() == "REVIEW":
        description = REVIEW_DESC
    elif skus_text.upper() == "ALL":
        description = ""
    else:
        description = describe(skus, labels)
        unknown = [s.strip() for s in skus_text.split("/") if s.strip() not in labels]
        if unknown:
            warnings.append("item %d: unknown SKU code(s) %s" % (item, ", ".join(unknown)))

    full_note = NOTE_PREFIX.get(confidence, "Inferred")
    if note:
        full_note += " - " + note

    lomak = ws.cell(row, COL_QTY_LOMAK + N_INSERTED).value or 0
    supplier = ws.cell(row, COL_QTY_LOMAK + N_INSERTED + 1).value or 0
    cost = unit_cost_usd(ws, row)
    if cost is None:
        warnings.append("item %d (%s): no unit cost in any currency - valued at 0" % (item, name))

    return Part(
        row=row,
        item=item,
        name=name,
        color=color,
        skus=skus,
        sku_text=skus_text,
        description=description,
        confidence=confidence,
        note=full_note,
        qty=float(lomak) + float(supplier),
        qty_per_unit=qty_per_unit,
        unit_cost_usd=cost,
    )


# --------------------------------------------------------------------------------
# workbook transform
# --------------------------------------------------------------------------------
def insert_columns(ws):
    """Insert 3 columns after the part name, keeping every formula pointing at the
    same data. openpyxl shifts neither formulas nor merged ranges, so do both here."""
    formulas = {
        (c.row, c.column): c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    }
    merges = [str(r) for r in ws.merged_cells.ranges]
    for rng in merges:
        ws.unmerge_cells(rng)

    ws.insert_cols(COL_COLOR, N_INSERTED)

    for (row, col), formula in formulas.items():
        new_col = col + N_INSERTED if col >= COL_COLOR else col
        origin = "%s%d" % (get_column_letter(col), row)
        target = "%s%d" % (get_column_letter(new_col), row)
        ws.cell(row, new_col).value = Translator(formula, origin=origin).translate_formula(target)

    for rng in merges:
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        shift = lambda c: c + N_INSERTED if c >= COL_COLOR else c  # noqa: E731
        ws.merge_cells(
            start_row=min_row,
            start_column=shift(min_col),
            end_row=max_row,
            end_column=shift(max_col),
        )


def write_annotations(ws, parts):
    for col, title in (
        (COL_COLOR, "color"),
        (COL_SKU, "SKU"),
        (COL_SKU_DESC, "SKU description"),
        (COL_NOTES, "Mapping basis / notes"),
    ):
        cell = ws.cell(HEADER_ROW, col, title)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    for part in parts:
        fill = {"inferred": FILL_INFERRED, "review": FILL_REVIEW}.get(part.confidence)
        for col, value in (
            (COL_COLOR, part.color),
            (COL_SKU, part.sku_text),
            (COL_SKU_DESC, part.description),
            (COL_NOTES, part.note),
        ):
            cell = ws.cell(part.row, col, value or None)
            cell.font = BODY_FONT
            if fill:
                cell.fill = fill

    for col, width in ((COL_COLOR, 14), (COL_SKU, 30), (COL_SKU_DESC, 58), (COL_NOTES, 62)):
        ws.column_dimensions[get_column_letter(col)].width = width


def write_legend(ws, start_row, family, labels, rules):
    def put(row, col, value, font=BODY_FONT):
        cell = ws.cell(row, col, value)
        cell.font = font

    row = start_row
    put(row, 2, "LEGEND - colour / SKU mapping", SECTION_FONT)
    row += 1
    put(row, 2, family.scope)
    row += 2

    for sku, label in labels.items():
        put(row, 2, sku)
        put(row, 3, label.rsplit(", ", 1)[-1])
        row += 1
    put(row, 2, "ALL")
    put(row, 3, "common part - used across all %d SKUs" % len(labels))
    row += 1
    put(row, 2, "REVIEW")
    put(row, 3, "allocation could not be determined from the plytix sheet - needs your confirmation")
    row += 2

    inferred = [r for r in rules if r.color and r.confidence != "verified"]
    if inferred:
        put(row, 2, "Interior / inferred colour codes (NOT sourced from plytix):", SECTION_FONT)
        row += 1
        for rule in inferred:
            put(row, 2, rule.color)
            put(row, 3, "%s -> %s" % (rule.note, rule.skus) if rule.note else "-> " + rule.skus)
            row += 1
        row += 1

    for line in family.extra_legend:
        put(row, 2, line)
        row += 1
    if family.extra_legend:
        row += 1

    put(row, 2, "Yellow fill = inferred.   Orange fill = needs review.   No fill = verified.")
    row += 1
    put(row, 2, "Quantities are Total Qty. (Lomak + supplier). Values converted to US$ at "
                "RMB %s / HK$ %s per US$." % (FX["rmb"], FX["hkd"]))


# --------------------------------------------------------------------------------
# buildable sheet
# --------------------------------------------------------------------------------
def buildable_units(parts):
    """Lowest number of units the given parts allow, and the part that limits it."""
    best = None
    limiter = ""
    for part in parts:
        if part.qty_per_unit is None:
            continue
        units = math.floor(part.qty / part.qty_per_unit)
        if best is None or units < best:
            best, limiter = units, part.name
    return best, limiter


def write_buildable_sheet(wb, parts, labels, sku_colors, warnings):
    ws = wb.create_sheet("SKU buildable")
    for col, width in (
        ("A", 62), ("B", 14), ("C", 58), ("D", 17), ("E", 34),
        ("F", 17), ("G", 34), ("H", 17), ("I", 15), ("J", 18),
    ):
        ws.column_dimensions[col].width = width

    ws["A1"] = "HOW MANY UNITS CAN STILL BE BUILT"
    ws["A1"].font = TITLE_FONT

    headers = [
        "SKU", "Colour", "SKU description",
        "Max units from its own parts", "Limiting own part",
        "Units allowed by shared / common stock", "Limiting shared / common part",
        "Buildable now", "Stranded US$ (own parts)", "Own-parts stock US$",
    ]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(3, col, title)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    common = [p for p in parts if p.sku_text.upper() == "ALL"]
    row = 4
    total_stranded = 0.0
    for sku, label in labels.items():
        exclusive = [p for p in parts if p.skus == [sku]]
        shared = [p for p in parts if sku in p.skus and len(p.skus) > 1]
        excl_units, excl_limiter = buildable_units(exclusive)
        pool_units, pool_limiter = buildable_units(shared + common)
        # No parts of its own means nothing to build, whatever the common pool holds.
        if excl_units is None:
            buildable_now = 0
        elif pool_units is None:
            buildable_now = excl_units
        else:
            buildable_now = min(excl_units, pool_units)

        # Stranded is measured against the SKU's OWN bottleneck, not the common one:
        # a cheap shared accessory running low does not strand this SKU's body parts.
        stranded = 0.0
        stock_value = 0.0
        for part in exclusive:
            cost = part.unit_cost_usd or 0.0
            stock_value += part.qty * cost
            consumed = (excl_units or 0) * (part.qty_per_unit or 1)
            stranded += max(part.qty - consumed, 0) * cost
        total_stranded += stranded

        values = [
            sku,
            sku_colors.get(sku, "no parts in this colour on the list"),
            label,
            excl_units if excl_units is not None else "no parts on list", excl_limiter,
            pool_units if pool_units is not None else "n/a", pool_limiter,
            buildable_now, round(stranded, 2), round(stock_value, 2),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.font = BODY_FONT
            if col in (9, 10):
                cell.number_format = "#,##0.00"
            elif col in (4, 6, 8) and isinstance(value, int):
                cell.number_format = "#,##0"
        if not excl_units:
            for col in range(1, 11):
                ws.cell(row, col).fill = FILL_REVIEW
        row += 1

    row += 1
    total_value = sum((p.unit_cost_usd or 0.0) * p.qty for p in parts)
    common_value = sum((p.unit_cost_usd or 0.0) * p.qty for p in common)
    review_value = sum(
        (p.unit_cost_usd or 0.0) * p.qty for p in parts if p.sku_text.upper() == "REVIEW"
    )
    for label, value in (
        ("Total leftover inventory, US$", total_value),
        ("...held in common parts (ALL)", common_value),
        ("...held in parts still needing review (REVIEW)", review_value),
        ("...stranded in SKU-specific parts above that SKU's own bottleneck", total_stranded),
    ):
        ws.cell(row, 1, label).font = BODY_FONT if label.startswith("...") else SECTION_FONT
        cell = ws.cell(row, 4, round(value, 2))
        cell.font = BODY_FONT
        cell.number_format = "#,##0.00"
        row += 1

    row += 1
    notes = [
        ("HOW TO READ THIS", SECTION_FONT),
        ("Max units from its own parts: lowest count across parts used by that SKU and nothing "
         "else. This is the number that matters for excess stock - those parts fit one product "
         "only, so anything above this count has nowhere to go.", BODY_FONT),
        ("Units allowed by shared / common stock: a CEILING, not a capacity. Common stock is one "
         "pool, so these per-SKU figures cannot all be achieved at the same time. A single cheap "
         "line running low (see the limiting part named) caps every SKU at once.", BODY_FONT),
        ("Buildable now: the lower of the two, i.e. what could be assembled without buying "
         "anything. Assumes 1 piece per unit unless the mapping CSV says otherwise.", BODY_FONT),
        ("Fasteners, tapes and papers are excluded from the minimums - pieces-per-unit is "
         "unknown for those (marked '?' in the overrides CSV).", BODY_FONT),
        ("Stranded US$: value of that SKU's own parts held beyond its OWN bottleneck. Measured "
         "against column D, not column H - a shared accessory running low does not strand a "
         "SKU's body parts, since that accessory can be re-bought.", BODY_FONT),
        ("Orange row = no parts of its own remain for that SKU, or the count is zero.", BODY_FONT),
    ]
    for text, font in notes:
        ws.cell(row, 1, text).font = font
        row += 1

    if warnings:
        row += 1
        ws.cell(row, 1, "WARNINGS FROM THIS RUN").font = SECTION_FONT
        row += 1
        for warning in warnings:
            ws.cell(row, 1, warning).font = BODY_FONT
            row += 1


def normalise_fonts(ws):
    """Force every cell in the sheet to Arial 9 black.

    Cells carried over from the source workbook keep whatever the original list used
    (mixed typefaces, 12pt titles, coloured text), so sweep the whole sheet at the end.
    Bold is preserved; nothing else is."""
    for row in ws.iter_rows():
        for cell in row:
            bold = bool(cell.font and cell.font.bold)
            try:
                cell.font = HEADER_FONT if bold else BODY_FONT
            except AttributeError:
                pass  # merged-range continuation cell - handled below


def set_default_font(wb):
    """Point the workbook's default font at Arial 9 black.

    Two kinds of cell never carry a style of their own: the continuation cells inside a
    merged block (openpyxl drops any style set on them when it writes the file) and any
    cell the user types into later. Both fall back to font 0, which in the source lists
    is 12pt Chinese. openpyxl exposes no public setter for it."""
    wb._fonts[0] = copy(BODY_FONT)


# --------------------------------------------------------------------------------
def process(family):
    labels = read_skus(family.sku_csv)
    rules = read_color_map(family.color_map_csv)
    overrides = read_overrides(family.overrides_csv)
    warnings = []

    # Colour token belonging to each SKU, from the verified single-SKU rules.
    sku_colors = {
        r.skus: r.color for r in rules if r.confidence == "verified" and r.skus in labels
    }

    wb = openpyxl.load_workbook(family.source)
    ws = wb.active

    data_rows = []
    row = FIRST_DATA_ROW
    while isinstance(ws.cell(row, COL_ITEM).value, int):
        data_rows.append(row)
        row += 1
    last_row = ws.max_row

    insert_columns(ws)
    parts = [
        resolve_part(ws, r, rules, overrides, labels, sku_colors, warnings) for r in data_rows
    ]
    write_annotations(ws, parts)
    write_legend(ws, last_row + 3, family, labels, rules)
    write_buildable_sheet(wb, parts, labels, sku_colors, warnings)

    set_default_font(wb)
    for sheet in wb.worksheets:
        normalise_fonts(sheet)

    OUTPUTS.mkdir(exist_ok=True)
    wb.save(family.output)

    counts = {c: sum(1 for p in parts if p.confidence == c) for c in NOTE_PREFIX}
    print("[%s] %d parts -> %s" % (family.key, len(parts), family.output.relative_to(ROOT)))
    print("      verified %d | inferred %d | needs review %d"
          % (counts["verified"], counts["inferred"], counts["review"]))
    for warning in warnings:
        print("      WARNING: " + warning)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--family", choices=sorted(FAMILIES), help="run one family only")
    args = parser.parse_args()

    for key in ([args.family] if args.family else sorted(FAMILIES)):
        process(FAMILIES[key])
    return 0


if __name__ == "__main__":
    sys.exit(main())
